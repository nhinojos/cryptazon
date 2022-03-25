## Necessary Libararies.
import requests # Request real-time crypotcurrency data from Coinbase.
from openpyxl import Workbook, load_workbook # Manipulating Excel Spreadsheet. https://openpyxl.readthedocs.io/en/stable/.
import yagmail # Simplified Gmail Delivery. https://github.com/kootenpv/yagmail.
import pickle # To serialize python objects. https://docs.python.org/3/library/pickle.html
from bs4 import BeautifulSoup # For parsing HTML: https://www.crummy.com/software/BeautifulSoup/bs4/doc/
from selenium import webdriver # Temporary driver to load HTML: https://selenium-python.readthedocs.io/
import time # Records date


## Retrieves real-time cryptocurrency data from Coinbase.
    # If display _types is True, then the currency 
    # types available will be displayed in the console. 
def currency_data(display_types=False):
    # Coinbase API call.
    data=requests.get(
            'https://api.coinbase.com/v2/exchange-rates'
            ).json()['data']['rates']
    # Displays cryptoccurrency keys neatly to user.
    if display_types:
        print('Here is a list of trackable cryptocurrencies:')
        crypto_options=''
        for i,crypto in enumerate(data.keys()):
            crypto_options+=crypto+(10-len(crypto))*' '
            if i%16==0:
                print(crypto_options)
                crypto_options=''
        return None
    else:
        return data


## Tracks products price history, 
## tabulates cryptocurrency data, 
## and emails users. 
class ProductTracker:
    # Initialize generates new object properties:
        # self.filename is the user-decided name for file writing.
        # self.link is the hyperlink to the Amazon product.
        # self.cryptocurrencies is the list of all target cryptos. 
        # self.sender is the email delivering notifications.
        # self.recipient is the email being sent notifcations.
        # self.password is the sender email's password.
        # self.driver_location is the file location of chrome driver. 
            # The default is the C:\ drive. 
    def __init__(
            self, filename, link, cryptocurrencies,
            email_recipient, email_sender,  
            threshold=None, password=None,
            driver_location="C:\chromedriver.exe"):

        # Establishing necessary variables
        self.filename=filename
        self.link=link
        self.cryptocurrencies=cryptocurrencies
        self.email_sender=email_sender
        self.email_recipient=email_recipient
        self.driver_location=driver_location

        # Generates new Excel workbook.
        workbook=Workbook() 
        workbook.active.title="Data"
        workbook.create_sheet("Analysis") 
        ws_data=workbook["Data"]
        # Appends inital row & coumn titles
        ws_data['B1'].value='USD'
        i=0
        for col in ws_data.iter_cols(min_row=1,max_row=1,min_col=3,
                                    max_col=2+len(self.cryptocurrencies)):
            for cell in col:
                cell.value=self.cryptocurrencies[i]
                i+=1
        ws_data['A2']="Threshold"
        ws_data.column_dimensions['A'].width=17
        workbook.save(filename+'.xlsx')
        
        # Appends threshold rates under column headers
        if threshold==None:
            threshold=self.get_price()
        self.update_data(threshold)
        
        # Registers delivering email and password,
        # if not done so already.
        if password!=None:
            yagmail.register(email_sender,password)

        return 

    # Updates row of data in "Data" worksheet
    def update_data(self,price=None):
        # Establishing initial variables.
        workbook=load_workbook(self.filename+".xlsx")
        ws_data=workbook["Data"]
        
        
        # row_target depends on wheter new data is being added,
        # or if inputted price is altering threshold row.
        if price==None:
            price=self.get_price()
            row_target=ws_data.max_row+1
            row_target_title=ws_data.cell(row_target,1)
            row_target_title.value=time.strftime("%d %b %y, %H:%M%p",
                                                 time.localtime())
        else:
            row_target=2
        
        # Lists exchange rates of price, along with price itself. 
        exchange_rates=[price]
        for crypto in self.cryptocurrencies:
            exchange_rates.append(float(currency_data()[crypto])*price)
    
        # Appends all exchange rates to the the targeted row.
        for row in ws_data.iter_rows(min_row=row_target, max_row=row_target,
                                    min_col=2, max_col=1+len(exchange_rates)):
            for i,cell in enumerate(row):
                cell.value=exchange_rates[i]
        
        # Saving to workbook. 
        workbook.save(self.filename+'.xlsx')
        return

    ## Retrieves price through webscraping.
    def get_price(self):
        # Executes Selenium Chrome driver. 
        driver=webdriver.Chrome(self.driver_location) 
        driver.get(self.link)
        
        # Writes entire HTML to a document
        with open("product_page.html", "w",encoding='utf-8') as f:
            f.write(driver.page_source)
        
        # Retrieves price from HTML document. 
        with open("product_page.html", "rb") as f:
            soup=BeautifulSoup(f,'lxml')
        price_usd=soup.find('span',class_='a-offscreen').text.replace('$','')
        return float(price_usd)


    ## Saves object and workbook to directory.
    def update_pickle(self,filename=None):
        # A new filename may be specified to generate new save instance.
        if filename==None:
            filename=self.filename

        # Saving pickled object.
        pickle_file=open(filename+'.pickle','wb')
        pickle.dump(self,pickle_file)
        return

    ## Delivers email notification to user.
    def email_notification(self, title, content):
        date=time.strftime("%d %b %y, %H:%M%p",time.localtime())
        title="Amazon-Crypto Tracker:"+self.filename+date
        yagmail.SMTP(self.sender).send(self.recipient,title,content)
        return None



        






