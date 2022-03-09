## Necessary Libararies.
import requests # Request real-time crypotcurrency data from Coinbase.
from openpyxl import Workbook # Manipulating Excel Spreadsheet. Documentation: https://openpyxl.readthedocs.io/en/stable/.
import yagmail #Simplified Gmail Delivery. Documentation: https://github.com/kootenpv/yagmail.
import pickle 
from bs4 import BeautifulSoup
from selenium import webdriver



## Track products price history, tabulate cryptocurrency data, and email users. 
class productTracker:
    # Initialize generates new object properties:
        # self.link is the hyperlink to the Amazon product.
        # self.crypto_target is the set() of all cryptocurrencies being tracked.
        # self.workbook is the excel workbook where price data is stored and analyzed.
        # self.sender is the email delivering notifications.
        # self.recipient is the email being sent notifcations.
    def __init__(self,name=None, link=None, cryptocurrencies=None,sender=None,recipient=None,password_registered=False,password=None):
        ## Asks for a name for file naming
        self.name=name
        while self.name==None:
            self.name=input('Please provide a name for this product:')

        ## Asks and validates product's web location 
        self.link=link
        while True:
            try:
                requests.get(self.link)
            except:
                print('Link not valid.')
            else:
                break
            
            print("Please paste the link to the Amazon product you would like to track:")
            self.link=input()
            for i in range(3):
                print("")


        ## Scrapes website data for product name if necessary
        if name==None:
            print('No name inputted, scraping link for name...')
            self.name='<name_on_website>'
            print(self.name)
            print("If this name looks correct, press Enter. Otherwise, type a new name below:")
            self.name=input()


        ## Determining what cryptocurrencies the user will track.
        # Asks and validates inputted cryptocurrenncies.
        self.cryptocurrencies=cryptocurrencies
        while True:
            #Conversts string to list to set.
            if self.cryptocurrencies!=None:
                if type(self.cryptocurrencies)==str: # Turns string to set.
                    self.cryptocurrencies.upper()
                    self.cryptocurrencies.replace(' ','')
                    self.cryptocurrencies=set(self.cryptocurrencies.split(','))
                
                if type(self.cryptocurrencies)==set: # Is this set valid?
                    if self.cryptocurrencies.issubset(self.currencyData().keys()):
                        break

                print('Error, inputted currency types are not available.')
            
            # Displaying cryptocurrency options. Then, asks for user input.
            else:
                print('Here are the list of cryptocurrencies available to track:')
                crypto_options=''
                for i,crypto in enumerate(self.currencyData().keys()):
                    crypto_options+=crypto+(10-len(crypto))*' '
                    if i%16==0:
                        print(crypto_options)
                        crypto_options=''
            
            print('Please input the abbreviated cryptocurrencies you would like to track, separated by commas:')
            self.cryptocurrencies=input()
            for i in range(3):
                print("")


        ## Generate new Excel workbook.
            # Includes two sheets, 'Data' and 'Analysis.'
        self.workbook=Workbook() 
        self.workbook.active.title="Data" 
        for i,crypto in enumerate(self.crypto_target):
            self.workbook["Data"]['A'+str(i+2)]=crypto
        self.workbook.create_sheet("Analysis")
        

        ## Registrating user emails.
        self.recipient=recipient
        self.sender=sender
        if None in [sender,recipient] or password_registered==False:
            print("Intializing Email registration.")
            print("If you would like to restart this process, type 'restart' into the command line.")
        
        while True:
            if recipient==None: # Validating recipient email.
                print('What email would you like to be nofitied?')
                self.recipient=input()
                self.recipient.replace(' ','')
                if self.recipient=='restart':
                    recipient=None
                    continue
            
            if sender==None: # Validating sender email.
                print("What email are you using to deliver notifications?")
                self.sender=input()
                self.sender.replace(' ','')
                if self.sender=='restart':
                    sender=None
                    continue

            if password_registered==False: #Validating and regisering password.
                print('password_registered:',password_registered)
                if password==None:
                    print("What is the delivering emails' password?")
                    password=input() # Password will not be saved in object instance.
                    if password=='restart':
                        continue
                
                yagmail.register(self.sender,password) # Stores username and password in Python's secure storage keyring library.
            
            break
        
        return


    ## Notifies user through email.
    def emailNotification(self, title, content):
        title=''
        yagmail.SMTP(self.sender).send(self.recipient,title,content)
        return


    ## Webscrapes Amazon for price data.
    def scrapePrice(self):
        # Setups temporary chrome driver.
        driver=webdriver.Chrome("C:\chromedriver.exe") # LOCATION MUST CHANGE TO WHEREVER YOU STORE YOUR SELENIUM DRIVER
        driver.get(self.link)
        # Downloads HTML form BeautifulSoup to read.
        with open("product_page.html", "w",encoding='utf-8') as f:
            f.write(driver.page_source)
        with open("product_page.html", "rb") as f:
            soup=BeautifulSoup(f,'lxml')
        # Returns the price within the HTML file.
        return soup.find('span',class_='a-offscreen').text.replace('$','')


    ## Returns real-time cryptocurrency data from Coinbase as .json().
    def currencyData(self):
        return requests.get('https://api.coinbase.com/v2/exchange-rates').json()['data']['rates']


    ## Saves workbook to directory as Excel document. 
    def saveToDirectory(self,object_name='product_tracker'): #Subject to change since it salves both object and excel data. 
        self.workbook.save(object_name+'.xlsx')
        pickle.dump(self,open(object_name+'.pickle','wb'))
        return
    

    ## Appends price data to excel workbook
    def appendPrice(self):
        price_usd=self.scrapePrice()
        ws_data=self.workbook["Data"]
        for row in ws_data.iter_rows(min_col=1, max_col=1):
            pass



        time=time.strftime("%d %b %y, %H:%M",time.gmtime())    

        



## Testing object by creating with product
bucket_hat={"Link":'https://www.amazon.com/VIVICMW-Breathable-Bordered-Outdoor-Fishing/dp/B07PP5X1R2/ref=sr_1_15?crid=BZPV1QMESPUD&keywords=bucket+hat&qid=1646709505&sprefix=bucket+hat%2Caps%2C123&sr=8-15'
            ,"Cryptocurrencies" : {'BTC','ETH'}
            ,"Sender":'kitten.in.superposition@gmail.com'
            ,"Recipient":'noah.hinojos@gmail.com'
                                    }

bucketHat=productTracker('bucket_hat',
                        bucket_hat['Link'],
                        bucket_hat['Cryptocurrencies'],
                        bucket_hat['Sender'],
                        bucket_hat['Recipient'],
                        True)



